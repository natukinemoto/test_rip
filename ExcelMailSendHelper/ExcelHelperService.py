import json
import os
import openpyxl
import calendar
from Constant import Const

class ExcelHelper:
    # excelに第二土曜日講習会情報を追加する
    def add_secondSaturday(self, fileName):
        self.excelYear = self.get_excelCell(fileName, "U3")
        self.excelMonth = self.get_excelCell(fileName, "W3")
        calendarMonth = self.get_calendarMonth(self.excelYear, self.excelMonth, 2, 5) # 第2週土曜日
        cellStart = "C" + str(7 + calendarMonth)
        cellend = "D" + str(7 + calendarMonth)
        cellStartBreak = "E" + str(7 + calendarMonth)
        cellEndBreak = "F" + str(7 + calendarMonth)
        cellStartActually = "J" + str(7 + calendarMonth)
        cellEndActually = "K" + str(7 + calendarMonth)
        cellBreak = "L" + str(7 + calendarMonth)
        cellState = "T" + str(7 + calendarMonth)
        cellSeminar = "U" + str(7 + calendarMonth)
        self.excelState = self.get_excelCell(fileName, cellState)
        if self.excelState != "休日出勤":
            self.input_excelCell(fileName, cellState, "休日出勤")
            self.input_excelCell(fileName, cellStart, "11:00")
            self.input_excelCell(fileName, cellend, "19:00")
            self.input_excelCell(fileName, cellStartBreak, "12:00")
            self.input_excelCell(fileName, cellEndBreak, "13:00")
            self.input_excelCell(fileName, cellStartActually, "11:00")
            self.input_excelCell(fileName, cellEndActually, "19:00")
            self.input_excelCell(fileName, cellBreak, "1:00")
            self.input_excelCell(fileName, cellSeminar, "【フロント講習会】")
        
    # excelをコピーする
    def save_excelCopy(self, fileName):
        # ブックを取得
        self.excelBook = openpyxl.load_workbook(fileName, keep_vba=True)
        self.save_CopySaveFile(self.excelBook)
    # 編集済みのexcelを保存する
    def save_excelComplete(self, fileName):
        # ブックを取得
        self.excelBook = openpyxl.load_workbook(fileName, keep_vba=True)
        # ブックからシートを取得
        self.save_CompleteSaveFile(self.excelBook)
        
    # エクセルの値を取得
    def get_excelCell(self, fileName, cell):
        # ブックを取得
        self.excelBook = openpyxl.load_workbook(fileName, keep_vba=True)
        # ブックからシートを取得
        self.excelSheet = self.excelBook[Const().get_excelHelperSheetName()]
        # セルを取得
        return self.excelSheet[cell].value
    # エクセルの値を書き込み(excelを更新)
    def input_excelCell(self, fileName, cell, excelValue):
        # ブックを取得
        self.excelBook = openpyxl.load_workbook(fileName, keep_vba=True)
        # ブックからシートを取得
        self.excelSheet = self.excelBook[Const().get_excelHelperSheetName()]
        # セルへ書き込み
        self.excelSheet[cell] = excelValue
        
        self.save_temporarySaveFile(self.excelBook, fileName)
    # カレンダー取得(第二土曜日)
    def get_calendarMonth(self, year, month, number, week):
        #self.calendarMonth = calendar.monthcalendar(year, month)
        #return self.calendarMonth[1][6]
        if number < 1 or week < 0 or week > 6:
            return Exception()
        firstWeek, n = calendar.monthrange(year, month)
        day = 7 * (number - 1) + (week - firstWeek) % 7 + 1
        
        return day if day <= n else Exception()
    # データ保存(コピー)
    def save_CopySaveFile(self, excelBook):
        dirName = os.getcwd()
        path = Const().get_excelHelperCopySavePath()
        save_CopySaveFilePath = os.path.join(dirName, *path)
        os.remove(save_CopySaveFilePath)
        excelBook.save(save_CopySaveFilePath)
    # 編集済みデータ保存
    def save_CompleteSaveFile(self, excelBook):
        dirName = os.getcwd()
        path = Const().get_excelHelperCompleteSavePath()
        save_CompleteSaveFilePath = os.path.join(dirName, *path)
        os.remove(save_CompleteSaveFilePath)
        excelBook.save(save_CompleteSaveFilePath)
    # 編集途中のデータ保存
    def save_temporarySaveFile(self, excelBook, filePath):
        excelBook.save(filePath)